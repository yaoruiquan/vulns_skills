#!/usr/bin/env python3
"""
更新漏洞汇总表

将提交的漏洞信息添加到汇总表 xlsx 文件中
"""

import argparse
import os
from datetime import datetime
from openpyxl import load_workbook, Workbook

# 导入配置模块
try:
    from config import get_summary_path
    DEFAULT_SUMMARY_PATH = get_summary_path()
except ImportError:
    DEFAULT_SUMMARY_PATH = "/Users/yao/Documents/网安- AI应用开发/监管上报/汇总表/漏洞汇总表.xlsx"


def ensure_dir(filepath):
    """确保目录存在"""
    dir_path = os.path.dirname(filepath)
    if dir_path and not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)


def create_summary_file(filepath):
    """创建新的汇总表文件"""
    ensure_dir(filepath)
    wb = Workbook()
    ws = wb.active
    ws.title = "漏洞汇总"

    # 设置表头
    headers = ["漏洞标题", "影响厂商", "漏洞编号", "提交人员", "上报CNVD编号", "上报CNNVD编号", "上报日期"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 设置列宽
    column_widths = [60, 20, 15, 15, 20, 20, 15]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width

    wb.save(filepath)
    print(f"创建汇总表: {filepath}")
    return wb


def find_row_by_das_id(ws, das_id):
    """根据 DAS-ID 查找行号，返回 None 表示未找到"""
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=3).value  # 漏洞编号在第3列
        if cell_value == das_id:
            return row
    return None


def update_summary(
    title: str,
    vendor: str,
    das_id: str,
    submitter: str,
    cnvd_id: str,
    cnnvd_id: str,
    date: str,
    filepath: str = None
):
    """
    更新汇总表

    Args:
        title: 漏洞标题
        vendor: 影响厂商
        das_id: 漏洞编号 (DAS-ID)
        submitter: 提交人员
        cnvd_id: 上报 CNVD 编号（可为空）
        cnnvd_id: 上报 CNNVD 编号
        date: 上报日期
        filepath: 汇总表文件路径
    """
    if filepath is None:
        filepath = DEFAULT_SUMMARY_PATH

    # 如果文件不存在，创建新文件
    if not os.path.exists(filepath):
        wb = create_summary_file(filepath)
        ws = wb.active
    else:
        wb = load_workbook(filepath)
        ws = wb.active

    # 查找是否已存在该漏洞
    existing_row = find_row_by_das_id(ws, das_id)

    if existing_row:
        # 更新现有行
        row = existing_row
        print(f"更新现有记录 (行 {row}): {das_id}")
    else:
        # 添加新行
        row = ws.max_row + 1
        print(f"添加新记录 (行 {row}): {das_id}")

    # 写入数据
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=2, value=vendor)
    ws.cell(row=row, column=3, value=das_id)
    ws.cell(row=row, column=4, value=submitter)
    ws.cell(row=row, column=5, value=cnvd_id if cnvd_id else "")
    ws.cell(row=row, column=6, value=cnnvd_id)
    ws.cell(row=row, column=7, value=date)

    # 保存文件
    wb.save(filepath)
    print(f"汇总表已更新: {filepath}")
    print(f"  漏洞标题: {title}")
    print(f"  影响厂商: {vendor}")
    print(f"  漏洞编号: {das_id}")
    print(f"  提交人员: {submitter}")
    print(f"  CNVD编号: {cnvd_id or '(空)'}")
    print(f"  CNNVD编号: {cnnvd_id}")
    print(f"  上报日期: {date}")


def main():
    parser = argparse.ArgumentParser(description="更新漏洞汇总表")
    parser.add_argument("--title", required=True, help="漏洞标题")
    parser.add_argument("--vendor", required=True, help="影响厂商")
    parser.add_argument("--das-id", required=True, help="漏洞编号 (DAS-ID)")
    parser.add_argument("--submitter", default="", help="提交人员")
    parser.add_argument("--cnvd-id", default="", help="上报 CNVD 编号")
    parser.add_argument("--cnnvd-id", required=True, help="上报 CNNVD 编号")
    parser.add_argument("--date", default=None, help="上报日期 (默认今天)")
    parser.add_argument("--filepath", default=None, help="汇总表文件路径")

    args = parser.parse_args()

    # 默认使用今天的日期
    if args.date is None:
        args.date = datetime.now().strftime("%Y-%m-%d")

    update_summary(
        title=args.title,
        vendor=args.vendor,
        das_id=args.das_id,
        submitter=args.submitter,
        cnvd_id=args.cnvd_id,
        cnnvd_id=args.cnnvd_id,
        date=args.date,
        filepath=args.filepath
    )


if __name__ == "__main__":
    main()